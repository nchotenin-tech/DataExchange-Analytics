"""แปลงผลรายงานเป็นไฟล์ CSV / Excel"""
from __future__ import annotations

import io
import re

import pandas as pd

SAFE_SHEET = re.compile(r"[\\/*?:\[\]]")


def table_to_frame(t: dict, multi_header: bool = False) -> pd.DataFrame:
    """แปลง table dict (จาก engine) เป็น DataFrame พร้อมหัวคอลัมน์ภาษาไทย

    multi_header=True  -> หัวตาราง 2 ชั้น (ใช้กับ Excel)
    multi_header=False -> รวมเป็นชั้นเดียว "กลุ่ม - ย่อย" (ใช้กับ CSV)
    """
    cols = [c for c in t["columns"] if "|รวมซี่" not in c]
    row_label = t.get("row_label") or "รายการ"

    data = {}
    header = []
    for c in cols:
        vals = [r.get(c) for r in t["rows"]]
        if c == "row":
            top, sub = row_label, ""
        elif "|" in c:
            top, sub = c.split("|", 1)
        else:
            top, sub = c, ""
        header.append((top, sub))
        data[c] = vals

    df = pd.DataFrame(data)
    if multi_header:
        df.columns = pd.MultiIndex.from_tuples(header)
    else:
        df.columns = [t if not s else f"{t} - {s}" for t, s in header]
    return df


def overview_to_frame(rows: list[dict], level: str, labels: dict) -> pd.DataFrame:
    cf = (labels or {}).get("caries_free_label", "ปราศจากฟันผุ")
    dmft = (labels or {}).get("dmft_label", "DMFT เฉลี่ย")
    unit = "อำเภอ" if level != "unit" else "หน่วยบริการ"
    return pd.DataFrame([{
        f"รหัส{unit}": r["code"],
        unit: r["name"],
        "เป้าหมาย (คน)": r["target"],
        "ตรวจ (คน)": r["examined"],
        "ยังไม่ได้ตรวจ (คน)": r["pending"],
        "%ได้รับการตรวจฟัน": r["examined_pct"],
        "ผ่านเกณฑ์คุณภาพ (คน)": r["qualified"],
        "%ผ่านเกณฑ์คุณภาพ": r["qualified_pct"],
        f"%{cf}": r["caries_free_pct"],
        dmft: r["dmft"],
    } for r in rows])


def info_frame(rep: dict) -> pd.DataFrame:
    s, sc = rep["summary"], rep["scope"]
    scope = " › ".join(x for x in [sc.get("pvname"), sc.get("ampname"), sc.get("hoscode")] if x)
    return pd.DataFrame({
        "รายการ": ["กลุ่มอายุ", "ขอบเขต", "เด็กในเขตรับผิดชอบ (คน)",
                   "ได้รับการตรวจฟัน (คน)", "ร้อยละได้รับการตรวจฟัน",
                   "ผ่านเกณฑ์คุณภาพ (คน)", "ร้อยละผ่านเกณฑ์คุณภาพ",
                   "จำนวนหน่วยบริการ", "จำนวนอำเภอ", "วันที่ตรวจแรกสุด", "วันที่ตรวจล่าสุด",
                   "วันที่ออกรายงาน"],
        "ค่า": [rep["profile"]["label"], scope or "ทุกพื้นที่ในข้อมูล", s.get("target"),
                s.get("examined"), s.get("examined_pct"), s.get("qualified"),
                s.get("qualified_pct"), s.get("hospitals"), s.get("districts"),
                s.get("date_min"), s.get("date_max"),
                pd.Timestamp.today().strftime("%Y-%m-%d %H:%M")],
    })


def find_table(rep: dict, no: str) -> dict | None:
    for t in rep["tables"]:
        if str(t["no"]) == str(no):
            return t
    return None


def csv_bytes(df: pd.DataFrame) -> bytes:
    """utf-8-sig เพื่อให้ Excel บน Windows อ่านภาษาไทยได้ถูก"""
    return df.to_csv(index=False).encode("utf-8-sig")


# คอลัมน์รหัสที่ต้องบังคับเป็น "ข้อความ" ใน Excel ไม่งั้นเลข 0 นำหน้าหาย
CODE_COLS = {"รหัสหน่วยบริการ", "PID", "เลขบัตรประชาชน", "typearea", "บ้านเลขที่",
             "รหัสอำเภอ", "รหัสหน่วยบริการ "}


def xlsx_bytes(sheets: list[tuple[str, pd.DataFrame, list[str]]],
               text_cols: set | None = None) -> bytes:
    """sheets = [(ชื่อชีต, DataFrame, [บรรทัดหมายเหตุ]), ...]

    เขียนด้วย openpyxl โดยตรง เพราะ pandas เขียนหัวตาราง 2 ชั้นแบบไม่มี index ไม่ได้
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill("solid", fgColor="E2F2FC")
    head_font = Font(bold=True, color="17364F", size=10.5)
    note_font = Font(italic=True, size=9, color="7C96AD")
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="F2F8FD")
    thin = Side(style="thin", color="C9E1F2")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name, df, notes in sheets:
        sheet = SAFE_SHEET.sub("-", str(name))[:31] or "Sheet"
        ws = wb.create_sheet(sheet)

        r = 1
        for line in notes:
            ws.cell(row=r, column=1, value=line).font = note_font
            r += 1
        if notes:
            r += 1

        multi = isinstance(df.columns, pd.MultiIndex)
        tops = [c[0] if multi else str(c) for c in df.columns]
        subs = [c[1] if multi else "" for c in df.columns]
        hrows = 2 if multi and any(subs) else 1
        head_start = r

        for i, (top, sub) in enumerate(zip(tops, subs), start=1):
            ws.cell(row=head_start, column=i, value=top)
            if hrows == 2:
                ws.cell(row=head_start + 1, column=i, value=sub or None)

        if hrows == 2:
            i = 1
            while i <= len(tops):
                j = i
                while j < len(tops) and tops[j] == tops[i - 1] and subs[j]:
                    j += 1
                if subs[i - 1] and j > i:      # กลุ่มเดียวกันหลายคอลัมน์ -> รวมแนวนอน
                    ws.merge_cells(start_row=head_start, start_column=i,
                                   end_row=head_start, end_column=j)
                elif not subs[i - 1]:          # ไม่มีหัวย่อย -> รวมแนวตั้ง
                    ws.merge_cells(start_row=head_start, start_column=i,
                                   end_row=head_start + 1, end_column=i)
                i = max(j, i + 1)

        for rr in range(head_start, head_start + hrows):
            for cc in range(1, len(df.columns) + 1):
                cell = ws.cell(row=rr, column=cc)
                cell.fill, cell.font, cell.alignment, cell.border = \
                    head_fill, head_font, center, box

        first_data = head_start + hrows
        for ri, (_, row) in enumerate(df.iterrows()):
            is_total = str(row.iloc[0]).strip() == "รวม"
            for ci, v in enumerate(row, start=1):
                cell = ws.cell(row=first_data + ri, column=ci,
                               value=None if pd.isna(v) else v)
                cell.border = box
                if is_total:
                    cell.font, cell.fill = total_font, total_fill
                if text_cols and tops[ci - 1] in text_cols:
                    cell.number_format = "@"
                elif isinstance(v, float):
                    cell.number_format = "0.00"

        for i in range(1, len(df.columns) + 1):
            texts = [str(tops[i - 1]), str(subs[i - 1])] + \
                    [str(v) for v in df.iloc[:200, i - 1]]
            ws.column_dimensions[get_column_letter(i)].width = \
                min(34, max(10, max(len(t) for t in texts) + 3))
        ws.freeze_panes = ws.cell(row=first_data, column=2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def insight_frame(rep: dict) -> pd.DataFrame:
    """รวมสรุปประเด็นของทุกตารางไว้ในชีตเดียว"""
    tone = {"good": "ดี", "warn": "เฝ้าระวัง", "bad": "ต้องแก้ไข", "info": "ข้อมูล"}
    rows = []
    for t in rep["tables"]:
        ins = t.get("insight") or {}
        head = f"ตารางที่ {t['no']} {t['title']}"
        for s in ins.get("summary", []):
            rows.append({"ตาราง": head, "ระดับ": "ภาพรวม", "รายการ": "",
                         "สถานะ": tone.get(s["tone"], ""), "ข้อสรุป": s["text"]})
        for s in ins.get("rows", []):
            rows.append({"ตาราง": head, "ระดับ": t.get("row_label", "รายการ"),
                         "รายการ": s["row"], "สถานะ": tone.get(s["tone"], ""),
                         "ข้อสรุป": s["text"]})
    return pd.DataFrame(rows or [{"ตาราง": "-", "ระดับ": "-", "รายการ": "-",
                                  "สถานะ": "-", "ข้อสรุป": "ไม่มีข้อสรุป"}])


def report_sheets(rep: dict) -> list[tuple[str, pd.DataFrame, list[str]]]:
    """ทุกตารางในรายงาน สำหรับส่งออกเป็น Excel หลายชีตในไฟล์เดียว"""
    out = [("ข้อมูลรายงาน", info_frame(rep), []),
           ("สรุปประเด็น", insight_frame(rep), [])]
    rows = rep["units"] if rep.get("level") == "unit" else rep["districts"]
    if rows:
        out.append(("ภาพรวมรายพื้นที่",
                    overview_to_frame(rows, rep.get("level"), rep.get("overview_labels")), []))
    for t in rep["tables"]:
        out.append((f"ตารางที่ {t['no']}", table_to_frame(t, multi_header=True),
                    [f"ตารางที่ {t['no']} {t['title']}"] + list(t.get("conditions") or [])))
    return out
